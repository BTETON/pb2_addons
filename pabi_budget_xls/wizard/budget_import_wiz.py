# -*- coding: utf-8 -*-
# © <YEAR(S)> <AUTHOR(S)>
# License AGPL-3.0 or later (http://www.gnu.org/licenses/agpl.html).
import sys
import cStringIO

import openpyxl

from openerp import tools
from openerp import models, fields, api, _
from openerp.exceptions import Warning as UserError


class BudgetImportWizard(models.TransientModel):
    _name = 'budget.import.wizard'

    input_file = fields.Binary('Template')
    datas_fname = fields.Char('Import File Name')

    @api.model
    def _update_cost_control_lines(self, workbook, budget):
        CostCtrl_Sheet = workbook.get_sheet_by_name('CostControl_1')
        max_row = CostCtrl_Sheet.max_row
        costcontrol = CostCtrl_Sheet.cell(row=6, column=2).value
        costcontrol_id =  self.env['cost.control'].search([('name', '=', tools.ustr(costcontrol))])
        activity_group = CostCtrl_Sheet.cell(row=5, column=2).value
        activity_group_id =  self.env['account.activity.group'].search([('name', '=', tools.ustr(activity_group))])
        costcontrol_vals = {'plan_id' : budget.id,
                     'cost_control_id': costcontrol_id.id}
        costcontrol_line = self.env['budget.plan.unit.cost.control'].create(costcontrol_vals)
        line_row = 11
        lines = {}
        lines_to_create = []
        line_ids=False
        for row in range(line_row, max_row):
            line_vals = {'activity_group_id': activity_group_id.id,
                         'cost_control_line_id': costcontrol_line.id}
            activity = CostCtrl_Sheet.cell(row=row, column=1).value
            if activity:
                activity_id = self.env['account.activity'].search([('name', '=', tools.ustr(activity))])
                line_vals.update({'activity_id': activity_id.id})
            name = CostCtrl_Sheet.cell(row=row, column=2).value
            if name:
                line_vals.update({'name': tools.ustr(name)})
            
            act_unit = CostCtrl_Sheet.cell(row=row, column=4).value
            if not act_unit:
                continue
            
            m1 = CostCtrl_Sheet.cell(row=row, column=9).value
            if m1:
                line_vals.update({'m1': m1})
            m2 = CostCtrl_Sheet.cell(row=row, column=10).value
            if m2:
                line_vals.update({'m2': m2})
            m3 = CostCtrl_Sheet.cell(row=row, column=11).value
            if m3:
                line_vals.update({'m3': m3})
            m4 = CostCtrl_Sheet.cell(row=row, column=12).value
            if m4:
                line_vals.update({'m4': m4})
            m5 = CostCtrl_Sheet.cell(row=row, column=13).value
            if m5:
                line_vals.update({'m5': m5})
            m6 = CostCtrl_Sheet.cell(row=row, column=14).value
            if m6:
                line_vals.update({'m6': m6})
            m7 = CostCtrl_Sheet.cell(row=row, column=15).value
            if m7:
                line_vals.update({'m7': m7})
            m8 = CostCtrl_Sheet.cell(row=row, column=16).value
            if m8:
                line_vals.update({'m8': m8})
            m9 = CostCtrl_Sheet.cell(row=row, column=17).value
            if m9:
                line_vals.update({'m9': m9})
            m10 = CostCtrl_Sheet.cell(row=row, column=18).value
            if m10:
                line_vals.update({'m10': m10})
            m11 = CostCtrl_Sheet.cell(row=row, column=19).value
            if m11:
                line_vals.update({'m11': m11})
            m12 = CostCtrl_Sheet.cell(row=row, column=20).value
            if m12:
                line_vals.update({'m12': m12})
            cc_line = self.env['budget.plan.unit.cost.control.line'].create(line_vals)
            if not line_ids:
                line_ids = cc_line
            else:
                line_ids += cc_line

    @api.multi
    def update_budget_prepare(self, budget_ids, template=None):
        if not template:
            raise UserError(_('Please add .xlsx template.'))
        if not budget_ids:
            raise UserError(_('No budget to import.'))

        imp_file = template.decode('base64')
        stream = cStringIO.StringIO(imp_file)

        try:
            workbook = openpyxl.load_workbook(stream)
        except IOError as e:
            raise UserError(_(e.strerror))
        except ValueError as e:
            raise UserError(_(e.strerror))
        except:
            e = sys.exc_info()[0]
            raise UserError(_('Wrong file format. Please enter .xlsx file.'))
        budgets = self.env['budget.plan.unit'].browse(budget_ids)
        for budget in budgets:
            if budget.state != 'draft':
                raise UserError(
                    _('You can update budget plan only in draft state!'))


            NonCostCtrl_Sheet = workbook.get_sheet_by_name('Non_CostControl')
            max_row = NonCostCtrl_Sheet.max_row
            vals = {}

            bg_id = NonCostCtrl_Sheet.cell(row=1, column=5).value
            if budget.id != bg_id:
                raise UserError(
                    _('Validation Error\n Please enter plan related file!')
                )
            fiscal_year = NonCostCtrl_Sheet.cell(row=1, column=2).value
            fiscal_year_id = self.env['account.fiscalyear'].search(
                [('name', '=', tools.ustr(fiscal_year))])
            if fiscal_year_id:
                vals.update({'fiscalyear_id': fiscal_year_id.id})

            org = NonCostCtrl_Sheet.cell(row=2, column=2).value
            org_id =\
                self.env['res.org'].search(
                    ['|',
                     ('code', '=', tools.ustr(org)),
                     ('name_short', '=', tools.ustr(org))])
#             if org_id:
#                 vals.update({'org_id': org_id.id})

            section = NonCostCtrl_Sheet.cell(row=3, column=2).value
            section_id = self.env['res.section'].search(
                ['|',
                 ('code', '=', tools.ustr(section)),
                 ('name_short', '=', tools.ustr(org))])
            if section_id:
                vals.update({'section_id': section_id.id})
            export_date = NonCostCtrl_Sheet.cell(row=4, column=2).value
            if export_date:
                vals.update({'date': export_date})

            responsible_by = NonCostCtrl_Sheet.cell(row=5, column=2).value
            responsible_by_id = self.env['res.users'].search(
                [('name', '=', tools.ustr(responsible_by))])
            if section_id:
                vals.update({'creating_user_id': responsible_by_id.id})


            self._update_cost_control_lines(workbook, budget)
            line_row = 11
            lines = {}
            lines_to_create = []
            for row in range(line_row, max_row):
                line_vals = {}
                line_vals.update({'section_id': section_id.id,
                                  'plan_id': budget.id,
                                  'org_id': org_id.id})
                ag_group = NonCostCtrl_Sheet.cell(row=row, column=4).value
                if not ag_group:
                    break
                ag_group_id = self.env['account.activity.group'].search(
                    [('name', '=', tools.ustr(ag_group))])
                if ag_group_id:
                    line_vals.update({'activity_group_id': ag_group_id.id})
                description = NonCostCtrl_Sheet.cell(row=row, column=5).value
                if description == '=FALSE()':
                    description = ''
                unit = NonCostCtrl_Sheet.cell(row=row, column=7).value or 0.0
                act_unitprice\
                    = NonCostCtrl_Sheet.cell(row=row, column=8).value or 0.0
                activity_unit =\
                    NonCostCtrl_Sheet.cell(row=row, column=9).value or 0.0
                line_vals.update({
                    'unit': unit,
                    'activity_unit_price': act_unitprice,
                    'activity_unit': activity_unit,
                    'description': description,
                })
                total_act_budget = unit * act_unitprice * activity_unit
                line_id = NonCostCtrl_Sheet.cell(row=row, column=26).value
                p = 0
                col = 11
                total_month_budget = 0.0
                while p != 13:
                    val = NonCostCtrl_Sheet.cell(row=row, column=col).value
                    if val and not isinstance(val, long):
                        raise UserError(
                            _('Please insert float value on\
                             row: %s - column: %s') % (row, col))
                    if val:
                        total_month_budget += val
                    line_vals.update({'m' + str(p): val})
                    if col == 23:
                        break
                    col += 1
                    p += 1
                if line_id:
                    lines.update({int(line_id): line_vals})
                else:
                    lines_to_create.append(line_vals)
                if (total_act_budget - total_month_budget) != 0.0:
                    raise UserError(_('Please verify budget lines total!'))
            for line in budget.plan_line_ids:
                if lines.get(line.id, False):
                    line.write(lines[line.id])
            for line in lines_to_create:
                self.env['budget.plan.unit.line'].create(line)
            budget.write(vals)
            attachement_id = self.env['ir.attachment'].create({
                'name': self.datas_fname,
                'datas': stream.getvalue().encode('base64'),
                'datas_fname': self.datas_fname,
                'res_model': 'budget.plan.unit',
                'res_id': budget.id,
            })
            self.env['budget.plan.history'].create({
                'user_id': self.env.user.id,
                'operation_date': fields.Datetime.now(),
                'operation_type': 'import',
                'plan_id': budget.id,
                'attachement_id': attachement_id.id
            })
            return {}

    @api.multi
    def import_budget(self):
        budget_ids = self.env.context.get('active_ids')
        result = self.update_budget_prepare(budget_ids, self.input_file)
        if result.get('messages'):
            msg = False
            for line in result['messages']:
                if not msg:
                    msg = line['message']
                else:
                    msg = msg + '\n' + line['message']
            if msg:
                raise UserError(_('%s') % (msg,))
